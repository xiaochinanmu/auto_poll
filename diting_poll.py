from openpyxl import Workbook
import os
from selenium import webdriver
from selenium.common.exceptions import NoSuchElementException
from selenium.webdriver.common.by import By
from lxml import etree
import time
print("hello")
option = webdriver.ChromeOptions()
# 是为了禁用 Chrome 的自动化标志，以避免被网站检测到使用了自动化工具。
option.add_experimental_option("excludeSwitches", ['enable-automation'])
# 是为了指定 Chrome 使用的用户数据目录和配置文件目录。这样可以使用存储在目录中的个人偏好设置和登录信息。
# option.add_argument(r'--user-data-dir=C:\Users\wintoo\AppData\Local\Google\Chrome\User Data')
# option.add_argument("--profile-directory=Default")
# 如果需要在无界面模式下执行操作
option.add_argument('--headless')
tu = webdriver.Chrome(options=option)
print("hello2")
# 通过 WebDriver 对象进行操作，例如访问网页等
# tu.get('http://bdata.wintoo.io/web/diting/v3/#/index')
tu.get('file:///D:/Desktop/diting/diting_area.html')
# wait = WebDriverWait(tu, 20)  # 设置最长等待时间为 10 秒
# 等待元素可点击
#   tu.find_element(by=By.XPATH, value='/html/body/div[3]/div/div[2]/ul/li[8]').click()
time.sleep(4)  # /html/body/div[2]/div[1]/div[1]/ul/li[8]/span
# tu.find_element(by=By.XPATH, value='//*[@id="app"]/section/aside/div/ul/li[3]').click()
# time.sleep(4)
# tu.find_element(by=By.XPATH, value='//*[@id="tab-DeviceStatus"]').click()  # 50006-49386=170
# time.sleep(14)
# tu.find_element(by=By.XPATH, value='/html/body/div[1]/section/section/main/div/div/form/div[1]/div[3]/div/button[1]').click()
# time.sleep(4)


file_name = 'diting_file.xlsx'
sheet_name = 'Sheet1'
if os.path.exists(file_name):
    os.remove(file_name)  # 删除已存在的文件
    # 创建工作簿对象
workbook = Workbook()

    # 获取默认的活动工作表
sheet = workbook.active
sheet.title = sheet_name
    # 设置表头
sheet['A1'] = '设备ID'
sheet['B1'] = '车牌号'
sheet['C1'] = '区域'
sheet['D1'] = '上传时间'
sheet['E1'] = '上传数量'
sheet['F1'] = '在线时长'
sheet['G1'] = '行驶里程'


def print_first_element(lst, name):
    if len(lst) > 0:
        print(lst[0],end=" ")
    else:
        print(f"{name}",end=" ")
next_page=True
while next_page:
    page_source = tu.page_source
    i=0
    while True:
        i=i+1
        tree = etree.HTML(page_source)
        device_id = tree.xpath(f'/html/body/div[1]/section/section/main/div/div/div/div[1]/div[2]/div[2]/div[2]/div[3]/table/tbody/tr[{i}]/td[1]/div/text()')
        license_plate = tree.xpath(f'/html/body/div[1]/section/section/main/div/div/div/div[1]/div[2]/div[2]/div[2]/div[3]/table/tbody/tr[{i}]/td[2]/div/text()')
        region = tree.xpath(f'/html/body/div[1]/section/section/main/div/div/div/div[1]/div[2]/div[2]/div[2]/div[3]/table/tbody/tr[{i}]/td[3]/div/text()')
        upload_time = tree.xpath(f'/html/body/div[1]/section/section/main/div/div/div/div[1]/div[2]/div[2]/div[2]/div[3]/table/tbody/tr[{i}]/td[4]/div/text()')
        upload_count = tree.xpath(f'/html/body/div[1]/section/section/main/div/div/div/div[1]/div[2]/div[2]/div[2]/div[3]/table/tbody/tr[{i}]/td[7]/div/button/span/text()')
        online_duration = tree.xpath(f'/html/body/div[1]/section/section/main/div/div/div/div[1]/div[2]/div[2]/div[2]/div[3]/table/tbody/tr[{i}]/td[10]/div/div/text()')
        driving_distance = tree.xpath(f'/html/body/div[1]/section/section/main/div/div/div/div[1]/div[2]/div[2]/div[2]/div[3]/table/tbody/tr[{i}]/td[11]/div/div/text()')


        if len(license_plate) > 0:
            print(license_plate[0])
        else:
            license_plate.append("无")

        if len(region) > 0:
            print(region[0])
        else:
            region.append("无")

        if len(upload_time) > 0:
            print(upload_time[0])
        else:
            upload_time.append("无")

        if len(upload_count) > 0:
            print(upload_count[0])
        else:
            upload_count.append("无")

        if len(online_duration) > 0:
            print(online_duration[0])
        else:
            online_duration.append("无")

        if len(driving_distance) > 0:
            print(driving_distance[0])
        else:
            driving_distance.append("无")



        if len(device_id)==0:
            break
        else:
            print(device_id[0],end=" ")

        print(license_plate[0],end=" ")
        print(region[0],end=" ")
        print(upload_time[0],end=" ")
        print(upload_count[0],end=" ")
        print(online_duration[0],end=" ")
        print(driving_distance[0],end=" ")


        # print_first_element(seri_num, "序列号")
        # print_first_element(license_plate, "无车牌")
        # print_first_element(region, "无区域")
        # print_first_element(upload_time, "无时间")
        # print_first_element(upload_count, "零上传")
        # print_first_element(online_duration, "无时长")
        # print_first_element(driving_distance, "无里程")
        last_row = sheet.max_row + 1
        sheet.cell(row=last_row, column=1, value=device_id[0])
        sheet.cell(row=last_row, column=2, value=license_plate[0])
        sheet.cell(row=last_row, column=3, value=region[0])
        sheet.cell(row=last_row, column=4, value=upload_time[0])
        sheet.cell(row=last_row, column=5, value=upload_count[0])
        sheet.cell(row=last_row, column=6, value=online_duration[0])
        sheet.cell(row=last_row, column=7, value=driving_distance[0])



    # 将后两列设置为数字类型
    # for row in sheet.iter_rows(min_row=2, min_col=3, max_col=4):
    #     for cell in row:
    #         # 将value属性赋值给自己，以便保留原有的值
    #         cell.value = cell.value
    #         # 设置数字格式
    #         cell.number_format = '0.00'

    # 将包含数字文本的单元格转换为数字
    # for row in sheet.iter_rows(min_row=2, min_col=3, max_col=4):
    #     for cell in row:
    #         # 判断单元格的数据类型
    #         if cell.data_type == 'n':
    #             # 如果是数值类型，则为Python数字类型
    #             cell_value = cell.value
    #         else:
    #             # 如果是字符串类型，则尝试将其转换为数字类型
    #             try:
    #                 cell_value = float(cell.value)
    #             except ValueError:
    #                 cell_value = None
    #
    #         # 将转换后的值赋值回单元格
    #         cell.value = cell_value

    workbook.save(file_name)
    time.sleep(2)
    try:
        # 使用 XPath 定位元素，并点击
        element = tu.find_element(By.XPATH, '/html/body/div[3]/div/div[2]/ul/li[8]')
        element.click()

    except NoSuchElementException as e:
        print("未找到目标元素:", e)

    finally:
        # 关闭浏览器
        tu.quit()

